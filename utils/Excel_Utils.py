import openpyxl
import os

def read_excel_data(file_path,sheet_path='Sheet1'):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_path]
    headers = [cell.value for cell in next(sheet.iter_rows(1,1))]

    data = []
    for row in sheet.iter_rows(2,values_only=True):
        data.append(dict(zip(headers,row)))
    
    return data


# print(read_excel_data(file_path=os.path.join(os.getcwd(),'../test data/Dummy TD.xlsx')))